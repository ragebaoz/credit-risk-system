#!/usr/bin/env python3
"""
生成信用评估可交互报告 Excel
包含：主结果、指标评分规则、城市明细
"""
import sys
import os
import json
from datetime import datetime
from typing import Dict, Any, List

sys.path.insert(0, "/Users/yuxuanyu/credit-risk-system")

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.formatting.rule import ColorScaleRule
import pandas as pd

from src.evaluation.client_eval import ClientEvaluator


def run_evaluation(name: str, credit_code: str, industry: str, internal_data: Dict[str, Any]) -> Dict[str, Any]:
    """运行单客户评估"""
    evaluator = ClientEvaluator(use_browser=False)
    return evaluator.evaluate(name, credit_code=credit_code, industry=industry,
                              internal_data=internal_data, save=False)


def style_header(ws, row_idx):
    """给表头添加样式"""
    for cell in ws[row_idx]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")


def create_main_sheet(wb, results: List[Dict[str, Any]]):
    """主结果 sheet — 去掉预付比例"""
    ws = wb.active
    ws.title = "评估结果汇总"

    headers = ["客户名称", "综合评分", "评分卡等级", "规则调整后等级",
               "建议账期(天)", "建议授信(万元)", "复核周期",
               "基础信用", "财务健康", "库存周转", "履约行为", "外部风险",
               "触发规则", "门店规模系数", "评估时间"]
    ws.append(headers)
    style_header(ws, 1)

    for r in results:
        store_count = r['raw_data'].get('dp_store_count', 0) or 0
        # 计算规模系数（与 scorecard 一致）
        from src.models.scorecard import ScorecardModel
        sc = ScorecardModel()
        scale_val = sc._get_store_scale_multiplier(store_count)
        scale = f"{scale_val:.3f}x"

        row = [
            r['client_name'],
            r['total_score'],
            r['base_grade'],
            r['risk_grade'],
            r['suggested_days'],
            round(r['suggested_credit'] / 10000, 2),
            r['review_cycle'],
            r['dimension_scores'].get('basic_credit', 0),
            r['dimension_scores'].get('financial_health', 0),
            r['dimension_scores'].get('inventory_turnover', 0),
            r['dimension_scores'].get('payment_behavior', 0),
            r['dimension_scores'].get('external_risk', 0),
            ", ".join([v['name'] for v in r.get('veto_rules', [])]) or "无",
            scale,
            r['evaluation_date']
        ]
        ws.append(row)

    # 列宽
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 30)
        ws.column_dimensions[column].width = adjusted_width


def create_indicator_sheet(wb, results: List[Dict[str, Any]]):
    """指标评分规则 sheet — 每家企业26项指标的原始值和得分"""
    ws = wb.create_sheet("指标评分规则")

    # 收集所有指标
    all_indicators = []
    dim_names = {}
    for r in results:
        for dim_key, dim_detail in r.get('dimension_details', {}).items():
            dim_names[dim_key] = dim_detail.get('name', dim_key)
            for ind_key, ind_data in dim_detail.get('indicators', {}).items():
                all_indicators.append({
                    'client': r['client_name'],
                    'dimension': dim_names.get(dim_key, dim_key),
                    'indicator_key': ind_key,
                    'indicator_name': ind_data.get('name', ind_key),
                    'raw_value': ind_data.get('value'),
                    'score': ind_data.get('score', 0),
                    'weight': ind_data.get('weight', 0),
                    'fallback': ind_data.get('fallback', False)
                })

    # 去重获取指标列表（按第一个客户）
    indicator_list = []
    seen = set()
    for ind in all_indicators:
        key = (ind['dimension'], ind['indicator_key'], ind['indicator_name'])
        if key not in seen:
            seen.add(key)
            indicator_list.append(ind)

    # 构建表头：维度 | 指标 | 权重 | 客户1原始值 | 客户1得分 | 客户2原始值 | 客户2得分 | ...
    clients = [r['client_name'] for r in results]
    headers = ["维度", "指标", "权重"]
    for c in clients:
        headers.append(f"{c}_原始值")
        headers.append(f"{c}_得分")
    ws.append(headers)
    style_header(ws, 1)

    # 按维度分组写入
    for ind in indicator_list:
        row = [ind['dimension'], ind['indicator_name'], f"{ind['weight']*100:.0f}%"]
        for c in clients:
            # 找到该客户该指标的数据
            found = None
            for a in all_indicators:
                if a['client'] == c and a['indicator_key'] == ind['indicator_key']:
                    found = a
                    break
            if found:
                raw = found['raw_value']
                if raw is None:
                    raw_str = "—"
                elif isinstance(raw, float):
                    raw_str = f"{raw:.4f}" if abs(raw) < 1 else f"{raw:.2f}"
                else:
                    raw_str = str(raw)
                if found['fallback']:
                    raw_str += " (缺省)"
                row.append(raw_str)
                row.append(round(found['score'], 2))
            else:
                row.append("—")
                row.append("—")
        ws.append(row)

    # 列宽
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 25)
        ws.column_dimensions[column].width = adjusted_width

    # 冻结首行
    ws.freeze_panes = "A2"


def create_city_sheet(wb, city_data: Dict[str, Dict[str, Any]]):
    """城市明细 sheet"""
    ws = wb.create_sheet("城市门店明细")

    headers = ["客户名称", "城市", "门店数"]
    ws.append(headers)
    style_header(ws, 1)

    for client_name, data in city_data.items():
        breakdown = data.get("city_breakdown", {})
        for city, count in breakdown.items():
            ws.append([client_name, city, count])

    # 列宽
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 20)
        ws.column_dimensions[column].width = adjusted_width


def main():
    # 1. 运行3家企业评估
    clients = [
        {
            'name': '名创优品',
            'credit_code': '91440101MA59J1K15C',
            'industry': '零售',
            'internal_data': {
                'dp_store_count': 1226,
                'dp_paused_count': 0,
                'dp_avg_rating': 4.2,
                'inventory_turnover_days': 45,
            }
        },
        {
            'name': '潮品挚尚',
            'credit_code': '91440101MA59J1K15D',
            'industry': '零售',
            'internal_data': {
                'dp_store_count': 56,
                'dp_paused_count': 0,
                'dp_avg_rating': 4.0,
                'inventory_turnover_days': 60,
            }
        },
        {
            'name': '力达动漫',
            'credit_code': '91440101MA59J1K15E',
            'industry': '零售',
            'internal_data': {
                'dp_store_count': 1,
                'dp_paused_count': 0,
                'dp_avg_rating': 3.5,
                'inventory_turnover_days': 90,
            }
        },
    ]

    results = []
    for c in clients:
        print(f"评估 {c['name']}...")
        r = run_evaluation(c['name'], c['credit_code'], c['industry'], c['internal_data'])
        results.append(r)
        print(f"  总分: {r['total_score']}, 基础: {r['base_grade']}, 最终: {r['risk_grade']}, 授信: {r['suggested_credit']}")

    # 2. 加载城市数据
    city_data = {}
    for fname, cname in [
        ('/tmp/dianping_mingchuang_8cities.json', '名创优品'),
        ('/tmp/dianping_chaopin_8cities.json', '潮品挚尚'),
    ]:
        try:
            with open(fname) as f:
                city_data[cname] = json.load(f)
        except Exception as e:
            print(f"警告: 无法加载 {fname}: {e}")

    # 力达动漫没有城市数据（只有1家）
    city_data['力达动漫'] = {"city_breakdown": {"未抓取": 1}}

    # 3. 生成Excel
    wb = Workbook()
    create_main_sheet(wb, results)
    create_indicator_sheet(wb, results)
    create_city_sheet(wb, city_data)

    output_path = os.path.expanduser("~/Desktop/信用评估报告_可交互.xlsx")
    wb.save(output_path)
    print(f"\n✅ 报告已保存: {output_path}")


if __name__ == "__main__":
    main()
