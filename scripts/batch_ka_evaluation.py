#!/usr/bin/env python3
"""
KA客户批量评估
以 KA客户 为一个评估单位，旗下多个主体数据汇总后评分
报告生成完全复用 generate_detailed_report.py 的6Sheet格式
"""
import sys
import os
import json
import subprocess
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from typing import Dict, Any, List

from src.evaluation.client_eval import ClientEvaluator
from generate_detailed_report import (
    create_raw_data_sheet, create_scoring_process_sheet,
    create_rules_sheet, create_credit_calc_sheet, create_city_sheet,
    style_header,
)
from openpyxl import Workbook


# ============================================================================
# 主流程
# ============================================================================
def run_subprocess(script_name: str):
    """运行独立子进程脚本，避免 asyncio 冲突"""
    script_path = os.path.join(os.path.dirname(__file__), script_name)
    result = subprocess.run(
        [sys.executable, script_path],
        capture_output=False,
        text=True,
    )
    return result.returncode == 0


def merge_by_ka(raw_results: List[Dict[str, Any]]) -> Dict[str, Dict[str, Any]]:
    ka_map: Dict[str, List[Dict]] = {}
    for r in raw_results:
        ka_map.setdefault(r['ka'], []).append(r)

    merged = {}
    for ka, entities in ka_map.items():
        m = {
            'ka_name': ka,
            'brands': list(set(e['brand'] for e in entities)),
            'companies': [e['company'] for e in entities],
            'entity_count': len(entities),
            'registered_capital': 0,
            'paid_in_capital': 0,
            'insured_count': 0,
            'established_years': 0,
            'lawsuit_count': 0,
            'contract_dispute_count': 0,
            'dishonest_records': 0,
            'executed_count': 0,
            'restriction_count': 0,
            'abnormal_records': 0,
            'penalty_records': 0,
            'negative_news': 0,
            'self_risk_count': 0,
            'around_risk_count': 0,
            'tianyancha_scores': [],
            'pledge_freeze': 0,
            'company_status': '存续',
        }
        min_established = None
        for e in entities:
            d = e.get('data', {})
            if not d:
                continue
            m['registered_capital'] += d.get('registered_capital', 0) or 0
            m['paid_in_capital'] += d.get('paid_in_capital', 0) or 0
            m['insured_count'] += d.get('insured_count', 0) or 0
            m['lawsuit_count'] = max(m['lawsuit_count'], d.get('lawsuit_count', 0) or 0)
            m['contract_dispute_count'] = max(m['contract_dispute_count'], d.get('contract_dispute_count', 0) or 0)
            m['dishonest_records'] = max(m['dishonest_records'], d.get('dishonest_records', 0) or 0)
            m['executed_count'] = max(m['executed_count'], d.get('executed_count', 0) or 0)
            m['restriction_count'] = max(m['restriction_count'], d.get('restriction_count', 0) or 0)
            m['abnormal_records'] = max(m['abnormal_records'], d.get('abnormal_records', 0) or 0)
            m['penalty_records'] = max(m['penalty_records'], d.get('penalty_records', 0) or 0)
            m['negative_news'] = max(m['negative_news'], d.get('negative_news', 0) or 0)
            m['self_risk_count'] = max(m['self_risk_count'], d.get('self_risk_count', 0) or 0)
            m['around_risk_count'] = max(m['around_risk_count'], d.get('around_risk_count', 0) or 0)
            m['pledge_freeze'] = max(m['pledge_freeze'], d.get('pledge_freeze', 0) or 0)
            score = d.get('tianyancha_score')
            if score is not None:
                m['tianyancha_scores'].append(score)
            est = d.get('established_years', 0) or 0
            if est > 0 and (min_established is None or est > min_established):
                min_established = est
        m['established_years'] = round(min_established, 1) if min_established else 0
        m['tianyancha_score'] = round(sum(m['tianyancha_scores']) / len(m['tianyancha_scores']), 1) if m['tianyancha_scores'] else None
        reg = m['registered_capital']
        paid = m['paid_in_capital']
        m['paid_in_capital_ratio'] = round(paid / reg, 2) if reg > 0 and paid > 0 else 0.0
        merged[ka] = m
    return merged


def evaluate_ka(ka_name: str, merged_data: Dict[str, Any], dp_data: Dict[str, Any]) -> Dict[str, Any]:
    evaluator = ClientEvaluator(use_browser=False)
    store_count = dp_data.get('total_store_count', 0) or 0
    paused_count = dp_data.get('total_paused_count', 0) or 0

    eval_data = {
        'established_years': merged_data['established_years'],
        'registered_capital': merged_data['registered_capital'],
        'paid_in_capital': merged_data['paid_in_capital'] if merged_data['paid_in_capital'] > 0 else None,
        'paid_in_capital_ratio': merged_data['paid_in_capital_ratio'],
        'abnormal_records': merged_data['abnormal_records'],
        'penalty_records': merged_data['penalty_records'],
        'company_status': merged_data['company_status'],
        'insured_count': merged_data['insured_count'] if merged_data['insured_count'] > 0 else None,
        'tianyancha_score': merged_data['tianyancha_score'],
        'self_risk_count': merged_data['self_risk_count'],
        'around_risk_count': merged_data['around_risk_count'],
        'lawsuit_count': merged_data['lawsuit_count'],
        'contract_dispute_count': merged_data['contract_dispute_count'],
        'dishonest_records': merged_data['dishonest_records'],
        'executed_count': merged_data['executed_count'],
        'restriction_count': merged_data['restriction_count'],
        'pledge_freeze': merged_data['pledge_freeze'],
        'negative_news': merged_data['negative_news'],
        'dp_store_count': store_count,
        'dp_paused_count': paused_count,
        'industry': '零售',
        'cooperation_years': 2,
    }

    from src.models.scorecard import ScorecardModel
    from src.models.rules import RulesEngine
    scorecard = ScorecardModel()
    rules_engine = RulesEngine()

    _, _, _, policy_adjust = rules_engine.evaluate(eval_data, 'A', 100)
    score_result = scorecard.calculate_score(eval_data, policy_adjust)
    final_grade, veto_rules, warnings, _ = rules_engine.evaluate(
        eval_data, score_result['risk_grade'], score_result['total_score']
    )

    if final_grade != score_result['risk_grade']:
        final_policy = scorecard.credit_policies.get(final_grade, scorecard.credit_policies['C'])
        score_result['suggested_days'] = final_policy['days']
        score_result['suggested_credit'] = final_policy['max_credit']
        score_result['advance_ratio'] = final_policy['advance_ratio']
        score_result['review_cycle'] = final_policy['review_cycle']
        score_result['policy_desc'] = final_policy['desc']

    return {
        'client_name': ka_name,
        'ka_name': ka_name,
        'brands': ', '.join(merged_data['brands']),
        'companies': merged_data['companies'],
        'entity_count': merged_data['entity_count'],
        'total_score': score_result['total_score'],
        'base_grade': score_result['risk_grade'],
        'risk_grade': final_grade,
        'suggested_days': score_result['suggested_days'],
        'suggested_credit': score_result['suggested_credit'],
        'advance_ratio': score_result['advance_ratio'],
        'review_cycle': score_result['review_cycle'],
        'policy_desc': score_result['policy_desc'],
        'dimension_scores': score_result['dimension_scores'],
        'dimension_details': score_result['dimension_details'],
        'veto_rules': veto_rules,
        'warnings': warnings,
        'raw_data': eval_data,
        'is_retail': score_result.get('is_retail', False),
        '_entities': [],
    }


def create_ka_summary_sheet(wb, results: List[Dict]):
    ws = wb.create_sheet("0-KA客户汇总", 0)
    headers = ["KA客户", "品牌", "主体数量", "注册资本(万)", "实缴资本(万)", "参保人数",
               "天眼评分", "诉讼", "经营异常", "负面舆情", "总分", "等级", "授信(万)", "账期(天)"]
    ws.append(headers)
    style_header(ws, 1)
    for r in results:
        m = r['raw_data']
        ws.append([
            r['ka_name'], r['brands'], r['entity_count'],
            round(m['registered_capital'] / 10000, 2),
            round(m['paid_in_capital'] / 10000, 2) if (m.get('paid_in_capital') or 0) > 0 else '—',
            m['insured_count'] if m.get('insured_count', 0) > 0 else '—',
            m['tianyancha_score'] if m.get('tianyancha_score') else '—',
            m['lawsuit_count'], m['abnormal_records'], m['negative_news'],
            r['total_score'], r['risk_grade'],
            round(r['suggested_credit'] / 10000, 2), r['suggested_days'],
        ])
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[column].width = min(max_length + 2, 25)


def create_entity_detail_sheet(wb, raw_results: List[Dict]):
    ws = wb.create_sheet("6-主体明细")
    headers = ["KA客户", "品牌", "公司主体", "成立年限", "注册资本(万)", "实缴资本(万)",
               "参保人数", "天眼评分", "诉讼", "经营异常", "处罚", "负面舆情"]
    ws.append(headers)
    style_header(ws, 1)
    for e in raw_results:
        d = e.get('data', {})
        ws.append([
            e['ka'], e['brand'], e['company'],
            d.get('established_years', '—'),
            round(d.get('registered_capital', 0) / 10000, 2) if d.get('registered_capital') else '—',
            round(d.get('paid_in_capital', 0) / 10000, 2) if d.get('paid_in_capital', 0) > 0 else '—',
            d.get('insured_count', '—'),
            d.get('tianyancha_score', '—'),
            d.get('lawsuit_count', 0),
            d.get('abnormal_records', 0),
            d.get('penalty_records', 0),
            d.get('negative_news', 0),
        ])
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[column].width = min(max_length + 2, 25)


def main():
    print("=" * 60)
    print("KA客户批量评估开始")
    print("=" * 60)

    # 1. 天眼查抓取（独立子进程）
    print("\n>>> 第1步: 天眼查抓取（独立进程）...")
    if not run_subprocess("fetch_ka_tianyancha.py"):
        print("天眼查抓取失败")
        return
    with open("/tmp/ka_tianyancha.json") as f:
        raw_results = json.load(f)

    # 2. 大众点评抓取（独立子进程）
    print("\n>>> 第2步: 大众点评抓取（独立进程）...")
    if not run_subprocess("fetch_ka_dianping.py"):
        print("大众点评抓取失败")
        return
    with open("/tmp/ka_dianping.json") as f:
        dp_data = json.load(f)

    # 3. 汇总
    print("\n>>> 第3步: 按KA客户汇总...")
    merged = merge_by_ka(raw_results)
    for ka, data in merged.items():
        print(f"  {ka}: {data['entity_count']}个主体 | 注册资本{data['registered_capital']/10000:,.0f}万 | 参保{data['insured_count']}人 | 天眼{data['tianyancha_score'] or 'N/A'}")

    # 4. 评估
    print("\n>>> 第4步: 评分...")
    eval_results = []
    for ka_name in merged:
        r = evaluate_ka(ka_name, merged[ka_name], dp_data.get(ka_name, {}))
        r['_entities'] = [e for e in raw_results if e['ka'] == ka_name]
        eval_results.append(r)
        print(f"  {ka_name}: {r['total_score']}分 | {r['risk_grade']}级 | 授信{r['suggested_credit']/10000:,.0f}万/{r['suggested_days']}天")

    # 5. 生成Excel
    print("\n>>> 第5步: 生成Excel报告...")
    wb = Workbook()
    create_raw_data_sheet(wb, eval_results)
    create_scoring_process_sheet(wb, eval_results)
    create_rules_sheet(wb, eval_results)
    create_credit_calc_sheet(wb, eval_results)
    city_data = {r['ka_name']: dp_data.get(r['ka_name'], {"city_breakdown": {"暂无数据": 0}}) for r in eval_results}
    create_city_sheet(wb, city_data)
    create_ka_summary_sheet(wb, eval_results)
    create_entity_detail_sheet(wb, raw_results)

    output_path = os.path.expanduser("~/Desktop/KA客户信用评估报告.xlsx")
    wb.save(output_path)
    print(f"\n✅ 报告已保存: {output_path}")


if __name__ == "__main__":
    main()
