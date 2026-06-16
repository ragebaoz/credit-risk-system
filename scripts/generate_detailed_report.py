#!/usr/bin/env python3
"""
生成信用评估数据与逻辑明细 Excel
展示完整的评价过程：原始值 -> 指标得分 -> 维度加权分 -> 总分 -> 规则引擎 -> 授信
"""
import sys
import os
import json
from datetime import datetime
from typing import Dict, Any, List

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from src.evaluation.client_eval import ClientEvaluator
from src.models.scorecard import ScorecardModel
from src.models.rules import RulesEngine


def run_evaluation(name: str, credit_code: str, industry: str, internal_data: Dict[str, Any]) -> Dict[str, Any]:
    evaluator = ClientEvaluator(use_browser=True)
    return evaluator.evaluate(name, credit_code=credit_code, industry=industry,
                              internal_data=internal_data, save=False)


def style_header(ws, row_idx, bg_color="4472C4"):
    for cell in ws[row_idx]:
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def style_subheader(ws, row_idx, bg_color="B4C7E7"):
    for cell in ws[row_idx]:
        cell.font = Font(bold=True, size=10)
        cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def thin_border():
    return Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )


def apply_border(ws, start_row, start_col, end_row, end_col):
    for r in range(start_row, end_row + 1):
        for c in range(start_col, end_col + 1):
            ws.cell(row=r, column=c).border = thin_border()


def create_raw_data_sheet(wb, results: List[Dict[str, Any]]):
    """Sheet1: 原始数据汇总"""
    ws = wb.active
    ws.title = "1-原始数据"

    # 按维度分组展示原始字段
    field_groups = [
        ("基础信用", ["established_years", "registered_capital", "paid_in_capital_ratio",
                      "paid_in_capital", "abnormal_records", "penalty_records", "company_status", "tianyancha_score"]),
        ("公司规模", ["insured_count", "affiliated_company_count", "affiliated_company_health"]),
        ("履约行为", ["on_time_rate", "avg_overdue_days", "max_overdue_amount",
                      "executed_count", "restriction_count", "cooperation_years"]),
        ("外部风险", ["lawsuit_count", "contract_dispute_count", "dishonest_records",
                      "pledge_freeze", "negative_news", "dp_store_count", "dp_pause_ratio"]),
    ]

    field_names = {
        "established_years": "成立年限(年)",
        "registered_capital": "注册资本(元)",
        "paid_in_capital": "实缴资本(元)",
        "paid_in_capital_ratio": "实缴比例",
        "abnormal_records": "经营异常记录(次)",
        "penalty_records": "行政处罚记录(次)",
        "tianyancha_score": "天眼查综合评分",
        "insured_count": "参保人数",
        "affiliated_company_count": "关联公司数量",
        "affiliated_company_health": "关联公司经营健康度",
        "debt_ratio": "资产负债率",
        "current_ratio": "流动比率",
        "revenue_growth": "营收增长率",
        "cash_flow": "经营现金流",
        "net_margin": "净利润率",
        "sudden_death_risk": "资金链断裂风险(0-3)",
        "self_risk_count": "自身风险数",
        "inventory_turnover_days": "库存周转天数",
        "recent_sales_trend": "近3月销售环比",
        "on_time_rate": "回款准时率",
        "avg_overdue_days": "平均逾期天数",
        "max_overdue_amount": "最大逾期金额占比",
        "executed_count": "历史被执行人次数",
        "restriction_count": "限制高消费次数",
        "cooperation_years": "合作年限(年)",
        "lawsuit_count": "被告诉讼数量",
        "contract_dispute_count": "买卖合同纠纷",
        "dishonest_records": "失信被执行人(次)",
        "pledge_freeze": "股权冻结/质押",
        "negative_news": "负面舆情(条)",
        "dp_store_count": "大众点评门店数",
        "dp_pause_ratio": "门店暂停营业比例",
    }

    headers = ["维度", "指标", "说明"] + [r['client_name'] for r in results]
    ws.append(headers)
    style_header(ws, 1)

    for dim_name, fields in field_groups:
        for i, f in enumerate(fields):
            row = [dim_name if i == 0 else "", field_names.get(f, f), ""]
            for r in results:
                val = r['raw_data'].get(f)
                if val is None:
                    row.append("—")
                elif isinstance(val, float):
                    row.append(f"{val:.4f}" if abs(val) < 1 else f"{val:.2f}")
                else:
                    row.append(val)
            ws.append(row)

    # 行业信息
    ws.append(["", "", ""])
    row = ["行业信息", "行业类型", ""]
    for r in results:
        row.append(r['raw_data'].get('industry', '—'))
    ws.append(row)

    row = ["", "是否零售客户", "dp_store_count > 5"]
    for r in results:
        row.append("是" if r.get('is_retail') else "否")
    ws.append(row)

    # 列宽
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 20
    for i in range(4, 4 + len(results)):
        ws.column_dimensions[get_column_letter(i)].width = 18


def create_scoring_process_sheet(wb, results: List[Dict[str, Any]]):
    """Sheet2: 指标打分过程 — 核心逻辑明细"""
    ws = wb.create_sheet("2-打分过程")

    # 表头结构：维度 | 指标 | 权重 | 打分规则说明 | 客户原始值 | 指标原始分 | 加权分
    # 每个客户占3列
    clients = [r['client_name'] for r in results]

    # 构建表头（多行）
    row1 = ["", "", "", ""] + [c for c in clients for _ in range(3)]
    ws.append(row1)

    row2 = ["维度", "指标", "权重", "打分规则说明"]
    for c in clients:
        row2.extend([f"{c}_原始值", f"{c}_原始分(0-100)", f"{c}_加权分"])
    ws.append(row2)
    style_header(ws, 2)

    # 合并客户名单元格
    for i, c in enumerate(clients):
        start_col = 5 + i * 3
        end_col = start_col + 2
        ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)
        ws.cell(row=1, column=start_col).value = c
        ws.cell(row=1, column=start_col).font = Font(bold=True, size=12, color="FFFFFF")
        ws.cell(row=1, column=start_col).fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
        ws.cell(row=1, column=start_col).alignment = Alignment(horizontal="center", vertical="center")

    # 收集所有指标及其打分规则
    scorecard = ScorecardModel()
    all_dims = scorecard.dimensions

    for dim_key, dim_cfg in all_dims.items():
        dim_name = dim_cfg['name']
        dim_weight = dim_cfg['weight']

        # 维度小标题行
        dim_row = [dim_name, f"【维度权重: {dim_weight*100:.0f}%】", "", ""]
        for r in results:
            dim_row.extend(["", "", ""])
        ws.append(dim_row)
        row_idx = ws.max_row
        for cell in ws[row_idx][:4]:
            cell.font = Font(bold=True, size=10, color="FFFFFF")
            cell.fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")

        for ind_key, ind_cfg in dim_cfg['indicators'].items():
            ind_name = ind_cfg['name']
            ind_weight = ind_cfg['weight']
            scoring = ind_cfg['scoring']

            # 构建打分规则说明
            rule_desc = []
            if isinstance(scoring, list):
                for item in scoring:
                    min_v = item.get('min')
                    max_v = item.get('max')
                    score = item['score']
                    desc = item.get('desc', '')
                    if min_v is None and max_v is not None:
                        cond = f"<{max_v}"
                    elif min_v is not None and max_v is None:
                        cond = f"≥{min_v}"
                    elif min_v is not None and max_v is not None:
                        cond = f"{min_v}~{max_v}"
                    else:
                        cond = "全部"
                    rule_desc.append(f"{cond} → {score}分 ({desc})")
            elif 'thresholds' in scoring:
                for thresh in scoring['thresholds']:
                    cond = thresh['condition']
                    score = thresh['score']
                    rule_desc.append(f"{cond} → {score}分")
            elif 'ranges' in scoring:
                for rng in scoring['ranges']:
                    min_v = rng.get('min', '-∞')
                    max_v = rng.get('max', '+∞')
                    score = rng['score']
                    rule_desc.append(f"{min_v}~{max_v} → {score}分")
            rule_str = "; ".join(rule_desc)

            # 每个客户的数据
            row = ["", ind_name, f"{ind_weight*100:.0f}%", rule_str]
            for r in results:
                dim_detail = r.get('dimension_details', {}).get(dim_key, {})
                indicators = dim_detail.get('indicators', {})
                ind_data = indicators.get(ind_key, {})

                raw_val = ind_data.get('value')
                score = ind_data.get('score', 0)
                weighted = score * ind_weight

                if raw_val is None:
                    raw_str = "—"
                elif isinstance(raw_val, float):
                    raw_str = f"{raw_val:.4f}" if abs(raw_val) < 1 else f"{raw_val:.2f}"
                else:
                    raw_str = str(raw_val)
                if ind_data.get('fallback'):
                    raw_str += " (未提供真实数据，使用系统默认分10)"

                row.extend([raw_str, round(score, 2), round(weighted, 2)])

            ws.append(row)

        # 维度合计行
        total_row = ["", "维度小计", "", ""]
        for r in results:
            dim_score = r['dimension_scores'].get(dim_key, 0)
            total_row.extend(["", "", dim_score])
        ws.append(total_row)
        row_idx = ws.max_row
        for cell in ws[row_idx]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")

    # 总分行
    ws.append([""] * 4)
    total_row = ["总分", "加权汇总", "", ""]
    for r in results:
        total_row.extend(["", "", r['total_score']])
    ws.append(total_row)
    row_idx = ws.max_row
    for cell in ws[row_idx]:
        cell.font = Font(bold=True, size=12, color="FFFFFF")
        cell.fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")

    # 列宽
    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 8
    ws.column_dimensions['D'].width = 35
    for i in range(5, 5 + len(clients) * 3):
        ws.column_dimensions[get_column_letter(i)].width = 16

    ws.freeze_panes = "E3"


def create_rules_sheet(wb, results: List[Dict[str, Any]]):
    """Sheet3: 规则引擎明细"""
    ws = wb.create_sheet("3-规则引擎")

    headers = ["客户名称", "评分卡等级", "评分卡总分", "规则名称", "规则类型",
               "触发条件", "是否触发", "动作", "结果等级", "规则说明"]
    ws.append(headers)
    style_header(ws, 1)

    rules_engine = RulesEngine()

    # 所有规则定义
    all_rules = []
    for rule in rules_engine.hard_rules.get('veto', []):
        all_rules.append(('否决规则', rule))
    for rule in rules_engine.hard_rules.get('head_enterprise', []):
        all_rules.append(('保底规则', rule))
    for rule in rules_engine.hard_rules.get('new_client', []):
        all_rules.append(('新客户规则', rule))
    for rule in rules_engine.hard_rules.get('warning', []):
        all_rules.append(('预警规则', rule))

    for r in results:
        base_grade = r['base_grade']
        total_score = r['total_score']
        raw_data = r['raw_data']

        # 预处理数据用于条件判断
        pdata = rules_engine._preprocess_data(raw_data)

        triggered_names = {v['name'] for v in r.get('veto_rules', [])}
        triggered_warnings = {w['name'] for w in r.get('warnings', [])}

        for rule_type, rule in all_rules:
            triggered = rules_engine._check_condition(rule['condition'], pdata)
            is_triggered = "✅ 是" if triggered else "❌ 否"

            # 判断该规则是否实际影响了结果
            actual_effect = ""
            if triggered:
                if rule['name'] in triggered_names:
                    actual_effect = f"等级调整至 {r['risk_grade']}"
                elif rule['name'] in triggered_warnings:
                    actual_effect = "触发预警"

            ws.append([
                r['client_name'],
                base_grade,
                total_score,
                rule['name'],
                rule_type,
                rule['condition'],
                is_triggered,
                rule.get('action', rule.get('level', '')),
                actual_effect,
                rule.get('desc', '')
            ])

        ws.append([""] * 10)

    # 列宽
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 40)
        ws.column_dimensions[column].width = adjusted_width


def create_credit_calc_sheet(wb, results: List[Dict[str, Any]]):
    """Sheet4: 授信计算过程"""
    ws = wb.create_sheet("4-授信计算")

    headers = ["客户名称", "评分卡等级", "规则调整后等级", "基础账期(天)",
               "基础授信(万元)", "门店数", "门店规模系数", "计算方式",
               "最终授信(万元)", "最终账期(天)", "备注"]
    ws.append(headers)
    style_header(ws, 1)

    scorecard = ScorecardModel()

    for r in results:
        base_grade = r['base_grade']
        final_grade = r['risk_grade']
        store_count = r['raw_data'].get('dp_store_count', 0) or 0
        scale = scorecard._get_store_scale_multiplier(store_count)

        # 基础政策
        base_policy = scorecard.credit_policies.get(base_grade, scorecard.credit_policies['C'])
        final_policy = scorecard.credit_policies.get(final_grade, scorecard.credit_policies['C'])

        # 计算过程说明
        if store_count < 10:
            calc_desc = f"门店{store_count}家 < 10家，系数=0，授信=0"
        elif store_count >= 1000:
            calc_desc = f"门店{store_count}家 ≥ 1000家，封顶系数=5.0"
        elif store_count < 50:
            calc_desc = f"({store_count}-10)/40 = {scale:.3f}"
        else:
            calc_desc = f"1.0 + ({store_count}-50)/950×4 = {scale:.3f}"

        # 如果规则引擎调整了等级，说明调整后的政策
        note = ""
        if base_grade != final_grade:
            note = f"规则引擎: {base_grade} → {final_grade}, 政策从 '{base_policy['desc']}' 变为 '{final_policy['desc']}'"

        row = [
            r['client_name'],
            base_grade,
            final_grade,
            final_policy['days'],
            round(final_policy['max_credit'] / 10000, 2),
            store_count,
            f"{scale:.3f}",
            calc_desc,
            round(r['suggested_credit'] / 10000, 2),
            r['suggested_days'],
            note
        ]
        ws.append(row)

    # 列宽
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width


def create_city_sheet(wb, city_data: Dict[str, Dict[str, Any]]):
    """Sheet5: 城市门店明细"""
    ws = wb.create_sheet("5-城市明细")

    headers = ["客户名称", "城市", "门店数"]
    ws.append(headers)
    style_header(ws, 1)

    for client_name, data in city_data.items():
        breakdown = data.get("city_breakdown", {})
        for city, count in breakdown.items():
            ws.append([client_name, city, count])

    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 20)
        ws.column_dimensions[column].width = adjusted_width


def main():
    clients = [
        {
            'name': '名创优品',
            'credit_code': '91440101MA59J1K15C',
            'industry': '零售',
            'internal_data': {
                'dp_store_count': 1226,
                'dp_paused_count': 0,
                'dp_avg_rating': 4.2,
                'inventory_turnover_days': 45,
            }
        },
        {
            'name': '浙江如苔文化',
            'credit_code': '91330784MA29Q4QT4K',
            'industry': '零售',
            'internal_data': {
                'dp_store_count': 300,
                'dp_paused_count': 0,
                'dp_avg_rating': 4.0,
                'inventory_turnover_days': 60,
            }
        },
        {
            'name': '潮品挚尚',
            'credit_code': '91440101MA59J1K15D',
            'industry': '零售',
            'internal_data': {
                'dp_store_count': 56,
                'dp_paused_count': 0,
                'dp_avg_rating': 4.0,
                'inventory_turnover_days': 60,
            }
        },
        {
            'name': '力达动漫',
            'credit_code': '91440101MA59J1K15E',
            'industry': '零售',
            'internal_data': {
                'dp_store_count': 1,
                'dp_paused_count': 0,
                'dp_avg_rating': 3.5,
                'inventory_turnover_days': 90,
            }
        },
    ]

    results = []
    for c in clients:
        print(f"评估 {c['name']}...")
        r = run_evaluation(c['name'], c['credit_code'], c['industry'], c['internal_data'])
        results.append(r)
        print(f"  总分: {r['total_score']}, 基础: {r['base_grade']}, 最终: {r['risk_grade']}")

    # 加载城市数据
    city_data = {}
    for fname, cname in [
        ('/tmp/dianping_mingchuang_8cities.json', '名创优品'),
        ('/tmp/dianping_rutai_8cities.json', '浙江如苔文化'),
        ('/tmp/dianping_chaopin_8cities.json', '潮品挚尚'),
    ]:
        try:
            with open(fname) as f:
                city_data[cname] = json.load(f)
        except Exception as e:
            print(f"警告: 无法加载 {fname}: {e}")
    city_data['力达动漫'] = {"city_breakdown": {"未抓取": 1}}

    # 生成Excel
    wb = Workbook()
    create_raw_data_sheet(wb, results)
    create_scoring_process_sheet(wb, results)
    create_rules_sheet(wb, results)
    create_credit_calc_sheet(wb, results)
    create_city_sheet(wb, city_data)

    output_path = os.path.expanduser("~/Desktop/信用评估_数据与逻辑明细.xlsx")
    wb.save(output_path)
    print(f"\n✅ 明细报告已保存: {output_path}")


if __name__ == "__main__":
    main()
