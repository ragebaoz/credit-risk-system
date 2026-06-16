#!/usr/bin/env python3
"""
卡游渠道客户信用评估报告生成器
使用真实数据（名创优品、潮品挚尚、力达动漫），按当前评分卡逻辑输出完整评估明细 Excel
"""
import sys
import os
from datetime import datetime

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from src.models.scorecard import ScorecardModel
from src.models.rules import RulesEngine
from src.utils.config_loader import get_weights_config, get_rules_config


# ===========================
# 三家客户真实数据
# ===========================
CLIENTS = [
    {
        "name": "名创优品",
        "credit_code": "91440101MA5AKFAH81",
        "industry": "toys_entertainment",
        "is_new": False,
        "cooperation_years": 3.0,
        "data": {
            "established_years": 8.6,
            "registered_capital": 146862400,
            "paid_in_capital": 139693000,
            "company_status": "存续",
            "abnormal_records": 0,
            "penalty_records": 0,
            "insured_count": 689,
            "affiliated_company_count": 185,
            "affiliated_company_health": 185,
            "around_risk_count": 185,
            "self_risk_count": 11,
            "dp_store_count": 1226,
            "dp_paused_count": 0,
            "lawsuit_count": 13,
            "contract_dispute_count": 0,
            "dishonest_records": 0,
            "executed_count": 0,
            "restriction_count": 0,
            "pledge_freeze": 0,
            "pledge_freeze_count": 0,
            "negative_news": 0,
            "tianyancha_score": 90,
            # 内部交易数据（已合作客户）
            "on_time_rate": 0.95,
            "avg_overdue_days": 3,
            "max_overdue_amount": 0.02,
        }
    },
    {
        "name": "浙江如苔文化",
        "credit_code": "91330421MA5AKFAH84",
        "industry": "toys_entertainment",
        "is_new": True,
        "cooperation_years": 0,
        "data": {
            "established_years": 5.5,
            "registered_capital": 10000000,
            "paid_in_capital": 0,
            "company_status": "存续",
            "abnormal_records": 0,
            "penalty_records": 0,
            "insured_count": 0,
            "affiliated_company_count": 0,
            "affiliated_company_health": 0,
            "around_risk_count": 0,
            "self_risk_count": 0,
            "dp_store_count": 300,
            "dp_paused_count": 0,
            "lawsuit_count": 0,
            "contract_dispute_count": 0,
            "dishonest_records": 0,
            "executed_count": 0,
            "restriction_count": 0,
            "pledge_freeze": 0,
            "pledge_freeze_count": 0,
            "negative_news": 0,
            "tianyancha_score": None,
            # 新客户：内部交易数据缺失
            "on_time_rate": None,
            "avg_overdue_days": None,
            "max_overdue_amount": None,
        }
    },
    {
        "name": "潮品挚尚",
        "credit_code": "91440101MA5AKFAH82",
        "industry": "toys_entertainment",
        "is_new": True,
        "cooperation_years": 0.2,
        "data": {
            "established_years": 11.8,
            "registered_capital": 1000000,
            "paid_in_capital": 0,
            "company_status": "存续",
            "abnormal_records": 0,
            "penalty_records": 0,
            "insured_count": 25,
            "affiliated_company_count": 1,
            "affiliated_company_health": 1,
            "around_risk_count": 1,
            "self_risk_count": 1,
            "dp_store_count": 56,
            "dp_paused_count": 0,
            "lawsuit_count": 3,
            "contract_dispute_count": 0,
            "dishonest_records": 0,
            "executed_count": 0,
            "restriction_count": 0,
            "pledge_freeze": 0,
            "pledge_freeze_count": 0,
            "negative_news": 0,
            "tianyancha_score": 58,
            # 新客户：内部交易数据缺失
            "on_time_rate": None,
            "avg_overdue_days": None,
            "max_overdue_amount": None,
        }
    },
    {
        "name": "力达动漫",
        "credit_code": "91440101MA5AKFAH83",
        "industry": "toys_entertainment",
        "is_new": True,
        "cooperation_years": 0,
        "data": {
            "established_years": 9.7,
            "registered_capital": 500000,
            "paid_in_capital": 0,
            "company_status": "存续",
            "abnormal_records": 0,
            "penalty_records": 0,
            "insured_count": 146,
            "affiliated_company_count": 8,
            "affiliated_company_health": 8,
            "around_risk_count": 8,
            "self_risk_count": 5,
            "dp_store_count": 1,
            "dp_paused_count": 0,
            "lawsuit_count": 1,
            "contract_dispute_count": 0,
            "dishonest_records": 0,
            "executed_count": 0,
            "restriction_count": 0,
            "pledge_freeze": 0,
            "pledge_freeze_count": 0,
            "negative_news": 0,
            "tianyancha_score": 53,
            # 新客户：内部交易数据缺失
            "on_time_rate": None,
            "avg_overdue_days": None,
            "max_overdue_amount": None,
        }
    },
]


def thin_border():
    return Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )


def apply_all_borders(ws, start_row, start_col, end_row, end_col):
    for r in range(start_row, end_row + 1):
        for c in range(start_col, end_col + 1):
            ws.cell(row=r, column=c).border = thin_border()


def header_style(cell, bg="4472C4"):
    cell.font = Font(bold=True, color="FFFFFF", size=11)
    cell.fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def subheader_style(cell, bg="B4C7E7"):
    cell.font = Font(bold=True, size=10)
    cell.fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def highlight_row(ws, row_idx, bg="FFF2CC"):
    for cell in ws[row_idx]:
        cell.fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
        cell.font = Font(bold=True)


def run_evaluations():
    """对三家客户执行评估"""
    scorecard = ScorecardModel()
    rules = RulesEngine()
    results = []

    for client in CLIENTS:
        raw_data = dict(client["data"])
        raw_data["industry"] = client["industry"]
        raw_data["cooperation_years"] = client["cooperation_years"]

        # 1. 规则引擎预处理（获取新客户策略调整）
        _, _, _, policy_adjust = rules.evaluate(raw_data, 'A', 100)

        # 2. 评分卡计算
        score_result = scorecard.calculate_score(raw_data, policy_adjust)

        # 3. 规则引擎最终评估
        final_grade, veto_rules, warnings, _ = rules.evaluate(
            raw_data, score_result['risk_grade'], score_result['total_score']
        )

        # 4. 如果规则引擎调整了等级，同步更新账期建议
        if final_grade != score_result['risk_grade']:
            final_policy = scorecard.credit_policies.get(final_grade, scorecard.credit_policies['C'])
            score_result['suggested_days'] = final_policy['days']
            score_result['suggested_credit'] = final_policy['max_credit']
            score_result['advance_ratio'] = final_policy['advance_ratio']
            score_result['review_cycle'] = final_policy['review_cycle']
            score_result['policy_desc'] = final_policy['desc']
            # 重新应用策略调整
            if 'days_reduce' in policy_adjust:
                score_result['suggested_days'] = max(0, score_result['suggested_days'] - policy_adjust['days_reduce'])
            if 'credit_reduce' in policy_adjust:
                score_result['suggested_credit'] = round(score_result['suggested_credit'] * policy_adjust['credit_reduce'])
            # 重新应用门店规模系数
            if score_result.get('is_retail'):
                store_count = raw_data.get('dp_store_count', 0) or 0
                scale_multiplier = scorecard._get_store_scale_multiplier(store_count)
                score_result['suggested_credit'] = round(score_result['suggested_credit'] * scale_multiplier)

        results.append({
            "client": client,
            "raw_data": raw_data,
            "score_result": score_result,
            "final_grade": final_grade,
            "veto_rules": veto_rules,
            "warnings": warnings,
            "policy_adjust": policy_adjust,
        })

    return results


def create_summary_sheet(wb, results):
    """Sheet1: 汇总结果"""
    ws = wb.active
    ws.title = "1-评估汇总"

    headers = ["客户名称", "统一社会信用代码", "客户类型", "合作年限",
               "评分卡总分", "基础等级", "规则调整后等级", "最终账期(天)",
               "基础授信(万元)", "门店规模系数", "最终授信(万元)", "预付比例",
               "复核周期", "管理措施", "触发的关键规则"]
    ws.append(headers)
    for cell in ws[1]:
        header_style(cell, "1F4E78")

    scorecard = ScorecardModel()
    for r in results:
        client = r["client"]
        sr = r["score_result"]
        store_count = r["raw_data"].get("dp_store_count", 0) or 0
        scale = scorecard._get_store_scale_multiplier(store_count)
        base_policy = scorecard.credit_policies.get(sr['risk_grade'], scorecard.credit_policies['C'])

        key_rules = []
        for v in r["veto_rules"]:
            key_rules.append(f"{v['name']}({v['action']})")
        for w in r["warnings"]:
            key_rules.append(f"{w['name']}({w.get('level','')})")

        row = [
            client["name"],
            client["credit_code"],
            "零售客户" if sr.get("is_retail") else "非零售客户",
            client["cooperation_years"],
            sr["total_score"],
            sr["risk_grade"],
            r["final_grade"],
            sr["suggested_days"],
            round(base_policy["max_credit"] / 10000, 2),
            f"{scale:.3f}",
            round(sr["suggested_credit"] / 10000, 2),
            f"{sr['advance_ratio']*100:.0f}%",
            sr["review_cycle"],
            sr["policy_desc"],
            "; ".join(key_rules) if key_rules else "无"
        ]
        ws.append(row)

    # 样式
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                max_len = max(max_len, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 3, 40)

    ws.freeze_panes = "A2"


def create_raw_data_sheet(wb, results):
    """Sheet2: 原始数据明细"""
    ws = wb.create_sheet("2-原始数据")

    groups = [
        ("基础信用", ["established_years", "registered_capital", "paid_in_capital",
                       "company_status", "abnormal_records", "penalty_records", "tianyancha_score"]),
        ("公司规模", ["insured_count", "affiliated_company_count", "affiliated_company_health"]),
        ("履约行为", ["on_time_rate", "avg_overdue_days", "max_overdue_amount",
                       "executed_count", "restriction_count", "cooperation_years"]),
        ("外部风险", ["lawsuit_count", "contract_dispute_count", "dishonest_records",
                       "pledge_freeze", "negative_news", "dp_store_count", "dp_pause_ratio",
                       "sudden_death_risk", "self_risk_count"]),
    ]

    field_names = {
        "established_years": "成立年限(年)",
        "registered_capital": "注册资本(元)",
        "paid_in_capital": "实缴资本(元)",
        "company_status": "工商存续状态",
        "abnormal_records": "经营异常记录(次)",
        "penalty_records": "行政处罚记录(次)",
        "tianyancha_score": "天眼查综合评分",
        "insured_count": "参保人数",
        "affiliated_company_count": "关联公司数量",
        "affiliated_company_health": "关联公司经营健康度",
        "sudden_death_risk": "资金链断裂风险(0-3)",
        "self_risk_count": "自身风险数",
        "on_time_rate": "回款准时率",
        "avg_overdue_days": "平均逾期天数",
        "max_overdue_amount": "最大逾期金额占比",
        "executed_count": "历史被执行人次数",
        "restriction_count": "限制高消费次数",
        "cooperation_years": "合作年限(年)",
        "lawsuit_count": "被告诉讼数量",
        "contract_dispute_count": "买卖合同纠纷",
        "dishonest_records": "失信被执行人(次)",
        "pledge_freeze": "股权冻结/质押",
        "negative_news": "负面舆情(条)",
        "dp_store_count": "大众点评门店数",
        "dp_pause_ratio": "门店暂停营业比例",
    }

    headers = ["维度", "指标", "说明"] + [r["client"]["name"] for r in results]
    ws.append(headers)
    for cell in ws[1]:
        header_style(cell, "4472C4")

    for dim_name, fields in groups:
        for i, f in enumerate(fields):
            row = [dim_name if i == 0 else "", field_names.get(f, f), ""]
            for r in results:
                val = r["raw_data"].get(f)
                if val is None:
                    row.append("—")
                elif isinstance(val, float):
                    if abs(val) < 1 and val != 0:
                        row.append(f"{val:.4f}")
                    else:
                        row.append(f"{val:.2f}")
                else:
                    row.append(val)
            ws.append(row)

    # 行业信息
    ws.append([""] * len(headers))
    row = ["行业信息", "行业类型", ""]
    for r in results:
        row.append(r["raw_data"].get("industry", "—"))
    ws.append(row)
    row = ["", "是否零售客户", "dp_store_count > 5"]
    for r in results:
        row.append("是" if r["score_result"].get("is_retail") else "否")
    ws.append(row)
    row = ["", "平台评分", "天眼查/爱企查/企查查"]
    for r in results:
        score = r["raw_data"].get("tianyancha_score") or r["raw_data"].get("aiqicha_score") or r["raw_data"].get("qcc_score")
        row.append(score if score is not None else "—")
    ws.append(row)

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 24
    ws.column_dimensions["C"].width = 22
    for i in range(4, 4 + len(results)):
        ws.column_dimensions[get_column_letter(i)].width = 20


def create_scoring_sheet(wb, results):
    """Sheet3: 评分过程明细"""
    ws = wb.create_sheet("3-评分过程")

    scorecard = ScorecardModel()
    clients = [r["client"]["name"] for r in results]

    # 多行表头
    row1 = ["", "", "", ""] + [c for c in clients for _ in range(3)]
    ws.append(row1)
    row2 = ["维度", "指标", "指标权重", "打分规则说明"]
    for c in clients:
        row2.extend([f"{c}_原始值", f"{c}_得分", f"{c}_加权分"])
    ws.append(row2)
    for cell in ws[2]:
        header_style(cell, "4472C4")

    # 合并客户名单元格
    for i, c in enumerate(clients):
        start_col = 5 + i * 3
        ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=start_col + 2)
        cell = ws.cell(row=1, column=start_col)
        cell.value = c
        cell.font = Font(bold=True, size=12, color="FFFFFF")
        cell.fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for dim_key, dim_cfg in scorecard.dimensions.items():
        dim_name = dim_cfg["name"]
        dim_weight = dim_cfg["weight"]

        # 维度标题行
        dim_row = [dim_name, f"【维度权重: {dim_weight*100:.0f}%】", "", ""]
        for _ in clients:
            dim_row.extend(["", "", ""])
        ws.append(dim_row)
        row_idx = ws.max_row
        for cell in ws[row_idx][:4]:
            cell.font = Font(bold=True, size=10, color="FFFFFF")
            cell.fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")

        for ind_key, ind_cfg in dim_cfg["indicators"].items():
            ind_name = ind_cfg["name"]
            ind_weight = ind_cfg["weight"]
            scoring = ind_cfg["scoring"]

            rule_desc = []
            for item in scoring:
                min_v = item.get("min")
                max_v = item.get("max")
                score = item["score"]
                desc = item.get("desc", "")
                if min_v is None and max_v is not None:
                    cond = f"<{max_v}"
                elif min_v is not None and max_v is None:
                    cond = f"≥{min_v}"
                elif min_v is not None and max_v is not None:
                    cond = f"{min_v}~{max_v}"
                else:
                    cond = "全部"
                rule_desc.append(f"{cond} → {score}分 ({desc})")
            rule_str = "; ".join(rule_desc)

            row = ["", ind_name, f"{ind_weight*100:.0f}%", rule_str]
            for r in results:
                dim_detail = r["score_result"]["dimension_details"].get(dim_key, {})
                indicators = dim_detail.get("indicators", {})
                ind_data = indicators.get(ind_key, {})

                raw_val = ind_data.get("value")
                score = ind_data.get("score", 0)
                weighted = score * ind_weight

                if raw_val is None:
                    raw_str = "—"
                elif isinstance(raw_val, float):
                    if abs(raw_val) < 1 and raw_val != 0:
                        raw_str = f"{raw_val:.4f}"
                    else:
                        raw_str = f"{raw_val:.2f}"
                else:
                    raw_str = str(raw_val)

                if ind_data.get("fallback"):
                    raw_str += " (缺省分10)"
                if ind_data.get("capped_by_new_client_rule"):
                    raw_str += f" (新客户封顶，原{ind_data.get('original_score', '')}分)"

                row.extend([raw_str, round(score, 2), round(weighted, 2)])
            ws.append(row)

        # 维度小计
        total_row = ["", "维度小计", "", ""]
        for r in results:
            dim_score = r["score_result"]["dimension_scores"].get(dim_key, 0)
            total_row.extend(["", "", dim_score])
        ws.append(total_row)
        for cell in ws[ws.max_row]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")

    # 总分行
    ws.append([""] * 4)
    total_row = ["总分", "加权汇总", "", ""]
    for r in results:
        total_row.extend(["", "", r["score_result"]["total_score"]])
    ws.append(total_row)
    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True, size=12, color="FFFFFF")
        cell.fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")

    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 38
    for i in range(5, 5 + len(clients) * 3):
        ws.column_dimensions[get_column_letter(i)].width = 16
    ws.freeze_panes = "E3"


def create_rules_sheet(wb, results):
    """Sheet4: 规则引擎明细"""
    ws = wb.create_sheet("4-规则引擎")

    rules_engine = RulesEngine()
    all_rules = []
    for rule in rules_engine.hard_rules.get("veto", []):
        all_rules.append(("否决规则", rule))
    for rule in rules_engine.hard_rules.get("head_enterprise", []):
        all_rules.append(("保底规则", rule))
    for rule in rules_engine.hard_rules.get("new_client", []):
        all_rules.append(("新客户规则", rule))
    for rule in rules_engine.hard_rules.get("warning", []):
        all_rules.append(("预警规则", rule))

    headers = ["客户名称", "评分卡等级", "评分卡总分", "规则名称", "规则类型",
               "触发条件", "是否触发", "动作/等级", "结果说明", "规则描述"]
    ws.append(headers)
    for cell in ws[1]:
        header_style(cell, "4472C4")

    for r in results:
        base_grade = r["score_result"]["risk_grade"]
        total_score = r["score_result"]["total_score"]
        raw_data = r["raw_data"]
        pdata = rules_engine._preprocess_data(raw_data)

        triggered_names = {v["name"] for v in r["veto_rules"]}
        triggered_warnings = {w["name"] for w in r["warnings"]}

        for rule_type, rule in all_rules:
            triggered = rules_engine._check_condition(rule["condition"], pdata)
            is_triggered = "✅ 触发" if triggered else "❌ 未触发"

            actual_effect = ""
            if triggered:
                if rule["name"] in triggered_names:
                    actual_effect = f"等级影响: {base_grade} → {r['final_grade']}"
                elif rule["name"] in triggered_warnings:
                    actual_effect = "触发预警"

            ws.append([
                r["client"]["name"],
                base_grade,
                total_score,
                rule["name"],
                rule_type,
                rule["condition"],
                is_triggered,
                rule.get("action", rule.get("level", "")),
                actual_effect,
                rule.get("desc", "")
            ])
        ws.append([""] * 10)

    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                max_len = max(max_len, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 2, 45)


def create_credit_sheet(wb, results):
    """Sheet5: 授信计算过程"""
    ws = wb.create_sheet("5-授信计算")

    headers = ["客户名称", "评分卡等级", "规则调整后等级", "基础账期(天)",
               "基础授信(万元)", "门店数", "门店规模系数", "计算说明",
               "策略调整", "最终授信(万元)", "最终账期(天)", "备注"]
    ws.append(headers)
    for cell in ws[1]:
        header_style(cell, "4472C4")

    scorecard = ScorecardModel()
    for r in results:
        base_grade = r["score_result"]["risk_grade"]
        final_grade = r["final_grade"]
        store_count = r["raw_data"].get("dp_store_count", 0) or 0
        scale = scorecard._get_store_scale_multiplier(store_count)
        base_policy = scorecard.credit_policies.get(base_grade, scorecard.credit_policies["C"])
        final_policy = scorecard.credit_policies.get(final_grade, scorecard.credit_policies["C"])

        if store_count < 10:
            calc_desc = f"门店{store_count}家 < 10家，系数=0"
        elif store_count >= 1000:
            calc_desc = f"门店{store_count}家 ≥ 1000家，封顶系数=5.0"
        elif store_count < 50:
            calc_desc = f"({store_count}-10)/40 = {scale:.3f}"
        else:
            calc_desc = f"1.0 + ({store_count}-50)/950×4 = {scale:.3f}"

        policy_adj_str = ""
        pa = r.get("policy_adjust", {})
        if pa:
            parts = []
            if "days_reduce" in pa:
                parts.append(f"账期减{pa['days_reduce']}天")
            if "credit_reduce" in pa:
                parts.append(f"额度×{pa['credit_reduce']}")
            if "payment_behavior_max" in pa:
                parts.append(f"履约行为封顶{pa['payment_behavior_max']}分")
            policy_adj_str = "; ".join(parts)

        note = ""
        if base_grade != final_grade:
            note = f"规则引擎调整: {base_grade} → {final_grade}"
        if r["client"]["is_new"]:
            note += "; 新客户限制已应用" if note else "新客户限制已应用"

        row = [
            r["client"]["name"],
            base_grade,
            final_grade,
            final_policy["days"],
            round(final_policy["max_credit"] / 10000, 2),
            store_count,
            f"{scale:.3f}",
            calc_desc,
            policy_adj_str,
            round(r["score_result"]["suggested_credit"] / 10000, 2),
            r["score_result"]["suggested_days"],
            note
        ]
        ws.append(row)

    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                max_len = max(max_len, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 2, 50)


def create_city_sheet(wb, results):
    """Sheet6: 大众点评城市门店明细"""
    ws = wb.create_sheet("6-城市门店明细")

    city_data = {
        "名创优品": {
            "广州": 215, "北京": 192, "深圳": 185, "上海": 183,
            "成都": 167, "郑州": 108, "武汉": 93, "杭州": 83
        },
        "潮品挚尚": {
            "深圳": 34, "上海": 7, "北京": 6, "杭州": 6,
            "广州": 2, "武汉": 1, "成都": 0, "郑州": 0
        },
        "力达动漫": {
            "未统计": 1
        }
    }

    headers = ["客户名称", "城市", "门店数", "占比", "备注"]
    ws.append(headers)
    for cell in ws[1]:
        header_style(cell, "4472C4")

    for client_name, cities in city_data.items():
        total = sum(cities.values())
        for i, (city, count) in enumerate(cities.items()):
            row = [client_name if i == 0 else "", city, count,
                   f"{count/total*100:.1f}%" if total > 0 else "—",
                   ""]
            ws.append(row)
        # 小计
        ws.append(["", "合计", total, "100%", ""])
        for cell in ws[ws.max_row]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
        ws.append([""] * 5)

    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                max_len = max(max_len, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 3, 20)


def main():
    print("=" * 60)
    print("卡游渠道客户信用评估报告生成")
    print("=" * 60)

    results = run_evaluations()

    # 打印控制台摘要
    for r in results:
        c = r["client"]
        sr = r["score_result"]
        print(f"\n📊 {c['name']}")
        print(f"   总分: {sr['total_score']} | 基础等级: {sr['risk_grade']} → 最终: {r['final_grade']}")
        print(f"   账期: {sr['suggested_days']}天 | 授信: ¥{sr['suggested_credit']:,.0f}")
        if r["veto_rules"]:
            for v in r["veto_rules"]:
                print(f"   ⚠️ {v['name']}: {v['desc']}")
        if r["warnings"]:
            for w in r["warnings"]:
                print(f"   🔔 [{w.get('level','')}] {w['name']}: {w['desc']}")

    # 生成 Excel
    wb = Workbook()
    create_summary_sheet(wb, results)
    create_raw_data_sheet(wb, results)
    create_scoring_sheet(wb, results)
    create_rules_sheet(wb, results)
    create_credit_sheet(wb, results)
    create_city_sheet(wb, results)

    output_path = os.path.expanduser("~/Desktop/卡游渠道客户信用评估报告.xlsx")
    wb.save(output_path)
    print(f"\n✅ 报告已保存: {output_path}")


if __name__ == "__main__":
    main()
